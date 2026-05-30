import io
from datetime import date
from typing import Optional

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from app.schemas import IncidentResponse
from app.schemas.equipment import EquipmentFailureResponse
from app.schemas.safety import SafetyViolationResponse

_DEJAVU_PATHS = {
    "DejaVuSans": "/usr/share/fonts/TTF/DejaVuSans.ttf",
    "DejaVuSans-Bold": "/usr/share/fonts/TTF/DejaVuSans-Bold.ttf",
}

_fonts_registered = False


def _register_fonts():
    global _fonts_registered
    if _fonts_registered:
        return
    for name, path in _DEJAVU_PATHS.items():
        try:
            pdfmetrics.registerFont(TTFont(name, path))
        except Exception:
            pass
    _fonts_registered = True


def _title_style():
    return ParagraphStyle(
        "RuTitle",
        fontName="DejaVuSans-Bold",
        fontSize=14,
        leading=18,
        spaceAfter=6,
    )


def _normal_style():
    return ParagraphStyle(
        "RuNormal",
        fontName="DejaVuSans",
        fontSize=10,
        leading=13,
    )


def generate_incidents_excel(incidents: list[IncidentResponse]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Травматизм"

    headers = ["ID", "Дата", "Цех", "Тип", "Тяжесть", "Дни нетрудоспособности", "Описание"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    severity_colors = {
        "Лёгкий": "92D050",
        "Средний": "FFC000",
        "Тяжёлый": "FF6600",
        "Смертельный": "FF0000",
    }

    for i, inc in enumerate(incidents, 2):
        ws.cell(row=i, column=1, value=inc.id)
        ws.cell(row=i, column=2, value=str(inc.date))
        ws.cell(row=i, column=3, value=inc.department_name or str(inc.department_id))
        ws.cell(row=i, column=4, value=inc.incident_type)
        cell_sev = ws.cell(row=i, column=5, value=inc.severity)
        if inc.severity in severity_colors:
            cell_sev.fill = PatternFill(
                start_color=severity_colors[inc.severity],
                end_color=severity_colors[inc.severity],
                fill_type="solid",
            )
        ws.cell(row=i, column=6, value=inc.days_lost)
        ws.cell(row=i, column=7, value=inc.description or "")

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_incidents_pdf(
    incidents: list[IncidentResponse],
    stats: Optional[dict] = None,
) -> bytes:
    _register_fonts()

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4))
    elements = []

    elements.append(Paragraph("Отчёт по производственному травматизму СИПР", _title_style()))
    elements.append(Spacer(1, 5 * mm))
    elements.append(Paragraph(f"Дата формирования: {date.today()}", _normal_style()))
    elements.append(Spacer(1, 5 * mm))

    if stats:
        stats_data = [
            ["Показатель", "Значение"],
            ["Всего случаев", str(stats.get("total_incidents", 0))],
            ["Дни нетрудоспособности", str(stats.get("total_days_lost", 0))],
            ["Коэффициент частоты", f"{stats.get('frequency_rate', 0):.2f}"],
            ["Коэффициент тяжести", f"{stats.get('severity_rate', 0):.2f}"],
        ]
        t = Table(stats_data)
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
                    ("FONTNAME", (0, 0), (-1, 0), "DejaVuSans-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                ]
            )
        )
        elements.append(t)
        elements.append(Spacer(1, 10 * mm))

    headers = ["Дата", "Цех", "Тип", "Тяжесть", "Дни"]
    data = [headers]
    for inc in incidents[:100]:
        data.append([str(inc.date), inc.department_name or str(inc.department_id), inc.incident_type, inc.severity, str(inc.days_lost)])

    t = Table(data)
    t.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("FONTNAME", (0, 0), (-1, -1), "DejaVuSans"),
                ("FONTNAME", (0, 0), (-1, 0), "DejaVuSans-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]
        )
    )
    elements.append(t)

    doc.build(elements)
    return buf.getvalue()
